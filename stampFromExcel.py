import openpyxl
import pythoncom
pythoncom.CoInitialize()
import win32com
import win32com.client
import PyPDF2


totalSDs = int((input('How many shop drawings are being reviewed? ')))
print(totalSDs)
numList = []


for i in range((totalSDs)):
        drawingNums = (input('Enter the Shop Drawing Number: '))
        if drawingNums == '':
            break
        else: 
            numList.append(drawingNums)
            
            

def xlsxToPdf(path):
    xlApp = win32com.client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(path)
    ws = books.Worksheets[0]
    ws.Visible = 1
    ws.ExportAsFixedFormat(0, path + '.pdf')

    
    
def stampWriter(numList):
 
    log = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\ShopDrawingLog.xlsx')
    logSheet = log.get_sheet_by_name('Log')
   
    for i in range(11, logSheet.get_highest_row()+1):
        
        if str(logSheet['A' + str(i)].value) in numList:
            
            if logSheet['F' + str(i)].value == 'NET':
                
                wb = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B2'].value = 'X'
                sheetPath = (r'C:\Users\wbartos\Documents\Pyscripts\Stamp\newstamp' + str(i))
                wb.save(sheetPath + '.xlsx')
                
                xlsxToPdf(sheetPath)
                
            elif logSheet['F' + str(i)].value == 'ET':
                
                wb = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B3'].value = 'X'
                sheetPath = (r'C:\Users\wbartos\Documents\Pyscripts\Stamp\newstamp' + str(i))
                wb.save(sheetPath + '.xlsx')
                
                xlsxToPdf(sheetPath)
                
            elif logSheet['F' + str(i)].value == 'E&C':
                
                wb = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B4'].value = 'X'
                sheetPath = (r'C:\Users\wbartos\Documents\Pyscripts\Stamp\newstamp' + str(i))
                wb.save(sheetPath + '.xlsx')
                
                xlsxToPdf(sheetPath)
                
            elif logSheet['F' + str(i)].value == 'R&R':
                
                wb = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B6'].value = 'X'
                sheetPath = (r'C:\Users\wbartos\Documents\Pyscripts\Stamp\newstamp' + str(i))
                wb.save(sheetPath + '.xlsx')
                
                xlsxToPdf(sheetPath)
                
            elif logSheet['F' + str(i)].value == 'REJ':
                
                wb = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B7'].value = 'X'
                sheetPath = (r'C:\Users\wbartos\Documents\Pyscripts\Stamp\newstamp' + str(i))
                wb.save(sheetPath + '.xlsx')
                
                xlsxToPdf(sheetPath)
                
                
def transmittalWriter(numList):
    
    pdfFileObj = open(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\transmittal\test.pdf')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    
    
    log = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\ShopDrawingLog.xlsx')
    logSheet = log.get_sheet_by_name('Log')
    
    wb = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\transmittal\transmittal.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
        
    for logrow in range(11, logSheet.get_highest_row()+1):
        
        if str(logSheet['A' + str(i)].value) in numList:  
            for transrow in range(11, transmittal.get_highest_row() +1):
                sheet['B4'].value = logSheet['A3'].value
                sheet['B5'].value = logSheet['A4'].value
                sheet['B6'].value = logSheet['A5'].value
                sheet['B'+str(transrow)].value = logSheet['A' + str(logrow)].value
                sheet['C'+str(transrow)].value = pdfReader.numPages
                sheet['E'+str(transrow)].value = logSheet['C' + str(logrow)].value
                sheet['D'+str(transrow)].value = logSheet['F' + str(logrow)].value
            wb.save(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\transmittal\transmittal1.xlsx')
        
    
            
            
transmittalWriter(numList)
    


    