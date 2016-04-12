#WRITTEN BY WILLIAM BARTOS IN PYTHON 3.5#



#THESE PACKAGES ARE NECESSARY FOR PYINSTALLER TO CONVERT TO .EXE
import six 
import packaging
import packaging.version
import packaging.specifiers
import packaging.requirements


import re
import openpyxl
import pythoncom
pythoncom.CoInitialize()
import win32com
import win32com.client
import PyPDF2
import os
import warnings
import logging
import colorama
from colorama import init, Fore, Back, Style

init()
warnings.filterwarnings('ignore')




shopDict = {'Project':'171',
            'Log':(r'ShopDrawingLog.xlsx'),
            'Stamp':(r'./Templates/Stamp.xlsx'),
            'Out': (os.path.abspath('OUT')),
            'In': (os.path.abspath('IN')),
            'Transmittal': (r'./Templates/transmittal.xlsx'),
            'Header':(r'./Templates/header.pdf')
            }
            
            
stampDict = {'NET':('B12'),
            'ET':('B13'),
            'E&C':('B14'),
            'R&R':('B16'),
            'REJ':('B17')
            }



while True:

    userInput = (input('How many shop drawings are being reviewed? '))
    totalSDs = 0
    
    if userInput == '':
          print(Fore.RED + Style.BRIGHT + '\n No empty inputs please')
          print(Style.RESET_ALL)
     
    elif re.match(r"^[0-9]*$", userInput):
         totalSDs=userInput
         break
     
    else: 
          print(Fore.RED + Style.BRIGHT + '\n Please only enter numbers or letters')
          print(Style.RESET_ALL)

sdCount =1
numList = []

while sdCount <= ((int(totalSDs))):
     drawingNums = (input('\nEnter the Shop Drawing Number: '))
     
     if drawingNums == '':
          print(Fore.RED + Style.BRIGHT + '\nNo empty inputs please')
          print(Style.RESET_ALL)
     
     elif re.match(r"^[A-Za-z0-9]*$", drawingNums):
         numList.append(drawingNums)
         print(Fore.CYAN + Style.BRIGHT + '\n' + '%d of ' %(sdCount) + str(totalSDs) +' entered')
         print(Style.RESET_ALL)
         sdCount +=1
     
     else: 
          print(Fore.RED + Style.BRIGHT + '\nPlease only enter numbers or letters')
          print(Style.RESET_ALL)
       
                 
def stampWriter(numList):
    
    log = openpyxl.load_workbook(shopDict['Log'])
    logSheet = log.get_sheet_by_name('Log')
    
    for i in range(11, logSheet.get_highest_row()+1):               
        if str(logSheet['A' + str(i)].value) in numList:
            if logSheet['F' + str(i)].value in stampDict:
                sdTitle = logSheet['C' + str(i)].value
                wb = openpyxl.load_workbook(shopDict['Stamp'])
                sdNo = str(logSheet['A' + str(i)].value)
                sdTitle = logSheet['C' + str(i)].value
                wb = openpyxl.load_workbook(shopDict['Stamp'])
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B9'].value = 'SD# ' + sdNo + ' - ' + sdTitle
                sheet['B29'].value = logSheet['I' + str(i)].value
                sheet[stampDict[logSheet['F' + str(i)].value]].value = '✓'
                sheetPath = (shopDict['Out'] + '\\newstamp' + str(i))
                transmittalPath = (shopDict['Out'] +'\\testout' + str(i) + '.pdf')
                submittalPath = logSheet['K'+str(i)].value
                
        
                wb.save(sheetPath + '.xlsx')
     
                pdfMerger(xlsxToPdf(sheetPath),submittalPath,transmittalPath)
                addHeader(transmittalPath, sdNo,sdTitle )
                os.remove(transmittalPath)
                os.remove(sheetPath + '.pdf')
                os.remove(sheetPath + '.xlsx')
                

def xlsxToPdf(path):
    xlApp = win32com.client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(path + '.xlsx')
    ws = books.Worksheets[0]
    ws.Visible = 1
    ws.ExportAsFixedFormat(0, path + '.pdf')
    stampPdf = path + '.pdf'
    books.Close(True)
    return stampPdf
     

def pdfMerger(stamp,submittal, path):
    
    stampFile = open(stamp, 'rb')
    submittalFile = open(submittal,'rb')  
    merger = PyPDF2.PdfFileMerger()
    
    try:
        merger.merge(position = 0, fileobj = stampFile) 
        merger.merge(position = 2, fileobj = submittalFile)
        merger.write(open(path, 'wb'))

    finally:
        stampFile.close()
        submittalFile.close()
        
def addHeader(path, sdNo, sdTitle):
    
    pdfNoHeader = open(path, 'rb') #opens the generated PDF(from xlsxToPdf)
    pdfReader = PyPDF2.PdfFileReader(pdfNoHeader) #creates an object out of the NoHeader PDF
    firstPage = pdfReader.getPage(0) #grabs the first page 
    pdfHeader= open(shopDict['Header'], 'rb')
    pdfHeaderReader = PyPDF2.PdfFileReader(pdfHeader) #opens the header template
    firstPage.mergePage(pdfHeaderReader.getPage(0)) #merges the first page of the NoHeader PDF with the header template
    pdfWriter = PyPDF2.PdfFileWriter() #creates a new PDF
    pdfWriter.addPage(firstPage) #adds the watermarked page to the first page of the new PDF
    
    for pageNum in range(1, pdfReader.numPages):
           pageObj = pdfReader.getPage(pageNum) #gets each page from the no header PDF
           pdfWriter.addPage(pageObj) #adds each page of the no header PDF to the new PDF
           

    resultPdfFile = open(shopDict['Out'] + '\\SD#' + str(sdNo) + '_' + str(sdTitle) + '.pdf', 'wb') #FINISHED PDF
    pdfWriter.write(resultPdfFile)
    print(resultPdfFile)
    pdfNoHeader.close()
    pdfHeader.close()
    
    print(Fore.CYAN + Style.BRIGHT +'\nStamp %s of %s generated' %(sdNo, totalSDs))
    print(Style.RESET_ALL)
        
 
       
      
stampWriter(numList)   


print(Fore.MAGENTA + Style.BRIGHT + '\n' + 'PRESS ENTER TO EXIT')
print(Style.RESET_ALL)    
input()  








    