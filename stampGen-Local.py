import openpyxl
import pythoncom
pythoncom.CoInitialize()
import win32com
import win32com.client
import PyPDF2



shopDict = {'Project':'171',
            'Log':(r'\ShopDrawingLog.xlsx'),
            'Stamp':(r'\Stamp2.xlsx'),
            'Out': (r'\OUT') }
            

            
stampDict = {'NET':('B2'),
            'ET':('B3'),
            'E&C':('B4'),
            'R&R':('B6'),
            'REJ':('B7')}

totalSDs = int((input('How many shop drawings are being reviewed?')))
numList = []
pdfList = []



for i in range((totalSDs)):
     drawingNums = (input('Enter the Shop Drawing Number: '))
     numList.append(drawingNums)

                 

def xlsxToPdf(path):
    xlApp = win32com.client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(path + '.xlsx')
    ws = books.Worksheets[0]
    ws.Visible = 1
    ws.ExportAsFixedFormat(0, path + '.pdf')
    stampPdf = path + '.pdf'
    return stampPdf

def pdfMerger(stamp,submittal, path):
    
    stampFile = open(stamp, 'rb')
    submittalFile = open(submittal,'rb')
    
    merger = PyPDF2.PdfFileMerger()
    
    try:
        merger.merge(position = 0, fileobj = stampFile)
        merger.merge(position = 2, fileobj = submittalFile)
        merger.write(open(path, 'wb'))
        #output.close()
    
    finally:
        stampFile.close()
        submittalFile.close()
            
    
def stampWriter(numList):
    
    log = openpyxl.load_workbook(shopDict['Log'])
    logSheet = log.get_sheet_by_name('Log')
    
    for i in range(11, logSheet.get_highest_row()+1):               
        if str(logSheet['A' + str(i)].value) in numList:
            if logSheet['F' + str(i)].value in stampDict:
    
                wb = openpyxl.load_workbook(shopDict['Stamp'])
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet[stampDict[logSheet['F' + str(i)].value]].value = 'X'
                sheetPath = (shopDict['Out'] + '\\newstamp' + str(i))
                transmittalPath = (shopDict['Out'] +'\\testout' + str(i) + '.pdf')
                submittalPath = logSheet['K'+str(i)].value
        
                wb.save(sheetPath + '.xlsx')
    
    
            pdfMerger(xlsxToPdf(sheetPath),submittalPath,transmittalPath)
                

def transmittalWriter(numList):
    
    wbpath = (r'C:\Users\wbartos\Documents\Pyscripts\Stamp\transmittal')
    wbpathext = wbpath + '.xlsx'
    
    pdfFileObj = open(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\transmittal\test.pdf', 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    
    log = openpyxl.load_workbook(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\ShopDrawingLog.xlsx')
    logSheet = log.get_sheet_by_name('Log')
    
    wb = openpyxl.load_workbook(wbpathext)
    sheet = wb.get_sheet_by_name('Sheet1')
    
        
    for logrow in range(11, logSheet.get_highest_row()+1):
        if str(logSheet['A' + str(logrow)].value) in numList:  
            #for transrow in range(11,logSheet.get_highest_row()+1):
            print('found')
            sheet['B4'].value = logSheet['A3'].value
            sheet['B5'].value = logSheet['A4'].value
            sheet['B6'].value = logSheet['A5'].value
            sheet['B'+str(logrow + 6)].value = logSheet['A' + str(logrow)].value
            sheet['C'+str(logrow + 6)].value = pdfReader.numPages
            sheet['E'+str(logrow + 6)].value = logSheet['C' + str(logrow)].value
            sheet['D'+str(logrow + 6)].value = logSheet['F' + str(logrow)].value
        wb.save(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\Transmittal\testout.xlsx')
        xlsxToPdf(wbpathext)
        
    
            
stampWriter(numList)   
#transmittalWriter(numList)
#xlsxToPdf(r'C:\Users\wbartos\Documents\Pyscripts\Stamp\OUT171\newstamp27')




    