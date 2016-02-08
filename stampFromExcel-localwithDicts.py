import openpyxl
#import pythoncom
#pythoncom.CoInitialize()
#import win32com
#import win32com.client
import PyPDF2



shopDict = {'Project':'153',
            'Log':(r'/Users/Will/Code/TestFolder/ShopDrawingLog(1).xlsx'),
            'Stamp':(r'/Users/Will/Code/TestFolder/stamp.xlsx'),
            'Out': (r'/Users/Will/Code/TestFolder/OUT') }
            
shopDict2 = {'Project':'143',
'Log':(r'/Users/Will/Code/TestFolder/ShopDrawingLog(1).xlsx'),
'Stamp':(r'/Users/Will/Code/TestFolder/stamp.xlsx'),
'Out': (r'/Users/Will/Code/TestFolder/OUT') }

topDict = {shopDict['Project']:shopDict, shopDict2['Project']:shopDict2};

print(topDict['143']);
            
topDict.push(shopDict.key, shopDict);

stampDict = {'NET':('B2'),
            'ET':('B3'),
            'E&C':('B4'),
            'R&R':('B5'),
            'REJ':('B6')}

totalSDs = int((input('How many shop drawings are being reviewed? ')))
print(totalSDs)
numList = []
pdfList = []


for i in range((totalSDs)):
        drawingNums = (input('Enter the Shop Drawing Number: '))
        if drawingNums == '':
            break
        else: 
            numList.append(drawingNums)
            
            

def xlsxToPdf(path):
    xlApp = win32com.client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(path)
    ws = books.Worksheets[0]
    ws.Visible = 1
    ws.ExportAsFixedFormat(0, path + '.pdf')
    stampPdf = path + '.pdf'
    return stampPdf

def pdfMerger(stamp,submittal, path):
    
    stampFile = open(stamp, 'rb')
    submittalFile = open(submittal,'rb')
    
    merger = PyPDF2.PdfFileMerger()
    
    try:
        merger.merge(position = 0, fileobj = stampFile)
        merger.merge(position = 2, fileobj = submittalFile)
        submittalFile.close()
        output = merger.write(open(path + '.pdf', 'wb'))
        output.close()
    
    finally:
        stampFile.close()
            
    
def stampWriter(numList):
 
    log = openpyxl.load_workbook(shopDict['Log'])
    logSheet = log.get_sheet_by_name('Log')
   
    for i in range(11, logSheet.get_highest_row()+1):
        
        if str(logSheet['A' + str(i)].value) in numList:
            
            if logSheet['F' + str(i)].value in stampDict:
                
                wb = openpyxl.load_workbook(shopDict['Stamp'])
                sheet = wb.get_sheet_by_name('Stamp')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet[stampDict[logSheet['F' + str(i)].value]].value = 'X'
                sheetPath = (shopDict['Out'] + '/newstamp' + str(i))
                submittalPath = logSheet['K'+str(i)].value
                wb.save(sheetPath + '.xlsx')
                
                #xlsxToPdf(sheetPath)
                #pdfMerger(xlsxToPdf(sheetPath),submittalPath,shopDict['Out'])
                

def transmittalWriter(numList):
    
    wbpath = (r'C:\Users\William\PythonScripts\stamps\transmittal\transmittal')
    wbpathext = wbpath + '.xlsx'
    
    pdfFileObj = open(r'C:\Users\William\PythonScripts\stamps\transmittal\test.pdf', 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    
    log = openpyxl.load_workbook(shopDict['Log'])
    logSheet = log.get_sheet_by_name('Log')
    
    wb = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\transmittal\transmittal.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    
        
    for logrow in range(11, logSheet.get_highest_row()+1):
        if str(logSheet['A' + str(logrow)].value) in numList:  
            #for transrow in range(11,logSheet.get_highest_row()+1):
            print('found')
            sheet['B4'].value = logSheet['A3'].value
            sheet['B5'].value = logSheet['A4'].value
            sheet['B6'].value = logSheet['A5'].value
            sheet['B'+str(logrow)].value = logSheet['A' + str(logrow)].value
            sheet['C'+str(logrow)].value = pdfReader.numPages
            sheet['E'+str(logrow)].value = logSheet['C' + str(logrow)].value
            sheet['D'+str(logrow)].value = logSheet['F' + str(logrow)].value
        wb.save(r'C:\Users\William\PythonScripts\stamps\transmittal\output.xlsx')
        
    
            
            
#transmittalWriter(numList)
stampWriter(numList)



    