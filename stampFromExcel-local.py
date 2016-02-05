import openpyxl
import pythoncom
pythoncom.CoInitialize()
import win32com
import win32com.client
import PyPDF2


totalSDs = int((input('How many shop drawings are being reviewed? ')))
print(totalSDs)
numList = []
pdfList = []


for i in range((totalSDs)):
        drawingNums = (input('Enter the Shop Drawing Number: '))
        if drawingNums == '':
            break
        else: 
            numList.append(drawingNums)
            
            

def xlsxToPdf(path):
    xlApp = win32com.client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(path)
    ws = books.Worksheets[0]
    ws.Visible = 1
    ws.ExportAsFixedFormat(0, path + '.pdf')
    stampPdf = path + '.pdf'
    return stampPdf

def pdfMerger(stamp,submittal):
    
    stampFile = open(stamp, 'rb')
    submittalFile = open(submittal,'rb')
    
    merger = PyPDF2.PdfFileMerger()
    
    try:
        merger.merge(position = 0, fileobj = stampFile)
        merger.merge(position = 2, fileobj = submittalFile)
        submittalFile.close()
        output = merger.write(open(submittal, 'wb'))
        output.close()
    
    finally:
        stampFile.close()
        
    
    

    
    
def stampWriter(numList):
 
    log = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\ShopDrawingLog.xlsx')
    logSheet = log.get_sheet_by_name('Log')
   
    for i in range(11, logSheet.get_highest_row()+1):
        
        if str(logSheet['A' + str(i)].value) in numList:
            
            if logSheet['F' + str(i)].value == 'NET':
                
                wb = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B2'].value = 'X'
                sheetPath = (r'C:\Users\William\PythonScripts\stamps\newstamp' + str(i))
                submittalPath = logSheet['K'+str(i)].value
                wb.save(sheetPath + '.xlsx')
                
                #xlsxToPdf(sheetPath)
                pdfMerger(xlsxToPdf(sheetPath),submittalPath)
                
            elif logSheet['F' + str(i)].value == 'ET':
                
                wb = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B3'].value = 'X'
                sheetPath = (r'C:\Users\William\PythonScripts\stamps\newstamp' + str(i))
                submittalPath = logSheet['K'+str(i)].value
                wb.save(sheetPath + '.xlsx')
                
                #xlsxToPdf(sheetPath)
                pdfMerger(xlsxToPdf(sheetPath),submittalPath)
                
            elif logSheet['F' + str(i)].value == 'E&C':
                
                wb = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B4'].value = 'X'
                sheetPath = (r'C:\Users\William\PythonScripts\stamps\newstamp' + str(i))
                submittalPath = logSheet['K'+str(i)].value
                wb.save(sheetPath + '.xlsx')
                
                #xlsxToPdf(sheetPath)
                pdfMerger(xlsxToPdf(sheetPath),submittalPath)
                
            elif logSheet['F' + str(i)].value == 'R&R':
                
                wb = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B6'].value = 'X'
                sheetPath = (r'C:\Users\William\PythonScripts\stamps\newstamp' + str(i))
                submittalPath = logSheet['K'+str(i)].value
                wb.save(sheetPath + '.xlsx')
                
                #xlsxToPdf(sheetPath)
                pdfMerger(xlsxToPdf(sheetPath),submittalPath)
                
            elif logSheet['F' + str(i)].value == 'REJ':
                
                wb = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\Stamp2.xlsx')
                sheet = wb.get_sheet_by_name('Sheet1')
                sheet['B1'].value = logSheet['C' + str(i)].value
                sheet['B7'].value = 'X'
                sheetPath = (r'C:\Users\William\PythonScripts\stamps\newstamp' + str(i))
                submittalPath = logSheet['K'+str(i)].value
                wb.save(sheetPath + '.xlsx')
                
                #xlsxToPdf(sheetPath)
                pdfMerger(xlsxToPdf(sheetPath),submittalPath)
                
                
def transmittalWriter(numList):
    
    wbpath = (r'C:\Users\William\PythonScripts\stamps\transmittal\transmittal')
    wbpathext = wbpath + '.xlsx'
    
    pdfFileObj = open(r'C:\Users\William\PythonScripts\stamps\transmittal\test.pdf', 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    
    log = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\ShopDrawingLog.xlsx')
    logSheet = log.get_sheet_by_name('Log')
    
    wb = openpyxl.load_workbook(r'C:\Users\William\PythonScripts\stamps\transmittal\transmittal.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    
        
    for logrow in range(11, logSheet.get_highest_row()+1):
        if str(logSheet['A' + str(logrow)].value) in numList:  
            #for transrow in range(11,logSheet.get_highest_row()+1):
            print('found')
            sheet['B4'].value = logSheet['A3'].value
            sheet['B5'].value = logSheet['A4'].value
            sheet['B6'].value = logSheet['A5'].value
            sheet['B'+str(logrow)].value = logSheet['A' + str(logrow)].value
            sheet['C'+str(logrow)].value = pdfReader.numPages
            sheet['E'+str(logrow)].value = logSheet['C' + str(logrow)].value
            sheet['D'+str(logrow)].value = logSheet['F' + str(logrow)].value
        wb.save(r'C:\Users\William\PythonScripts\stamps\transmittal\output.xlsx')
        
    
            
            
#transmittalWriter(numList)
stampWriter(numList)



    